"""
Excel export utility for BRD projects.
COMPLETELY FIXED VERSION with proper data export to all sheets.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
import os

logger = logging.getLogger(__name__)

def export_to_excel(brd_project):
    """Export BRD project to multi-sheet Excel file."""
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define styles
        header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create sheets
        create_overview_sheet(wb, brd_project, header_fill, header_font, border)
        create_ui_spec_sheet(wb, brd_project, header_fill, header_font, border)
        create_api_spec_sheet(wb, brd_project, header_fill, header_font, border)
        create_llm_prompt_sheet(wb, brd_project, header_fill, header_font, border)
        create_database_schema_sheet(wb, brd_project, header_fill, header_font, border)
        create_tech_stack_sheet(wb, brd_project, header_fill, header_font, border)
        create_traceability_sheet(wb, brd_project, header_fill, header_font, border)
        
        # Add agent sheets if they have data
        if hasattr(brd_project, 'agent_architectures') and brd_project.agent_architectures:
            create_agent_architecture_sheet(wb, brd_project, header_fill, header_font, border)
        if hasattr(brd_project, 'agent_configurations') and brd_project.agent_configurations:
            create_agent_configuration_sheet(wb, brd_project, header_fill, header_font, border)
        if hasattr(brd_project, 'agent_tasks') and brd_project.agent_tasks:
            create_agent_task_sheet(wb, brd_project, header_fill, header_font, border)
        
        # Save file
        os.makedirs("exports", exist_ok=True)
        filename = f"exports/BRD_{brd_project.overview.project_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        logger.info(f"Excel file exported: {filename}")
        return filename
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        raise


def create_overview_sheet(wb, brd_project, header_fill, header_font, border):
    """Create Overview sheet."""
    ws = wb.create_sheet("1. Overview", 0)
    
    # Headers
    headers = ["Field", "Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    overview = brd_project.overview
    data = [
        ["Project Name", overview.project_name or ""],
        ["Project Description", overview.project_description or ""],
        ["Business Goal", overview.business_goal or ""],
        ["Document Version", overview.document_version or ""],
        ["Prepared By", overview.prepared_by or ""],
        ["Approved By", overview.approved_by or ""],
        ["Target Release Date", overview.target_release_date or ""],
        ["Created At", str(datetime.now())],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def create_ui_spec_sheet(wb, brd_project, header_fill, header_font, border):
    """Create UI Specification sheet."""
    ws = wb.create_sheet("2. UI Specification", 1)
    
    # Headers
    headers = ["Screen/Component", "Requirement Description", "Business Rule", "Requirement ID", "Feature/Module"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    for row_idx, spec in enumerate(brd_project.ui_specifications, 2):
        row_data = [
            spec.screen_component or "",
            spec.requirement_description or "",
            spec.business_rule or "",
            spec.requirement_id or "",
            spec.feature_module or ""
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20


def create_api_spec_sheet(wb, brd_project, header_fill, header_font, border):
    """Create API Specification sheet."""
    ws = wb.create_sheet("3. API Specification", 2)
    
    # Headers
    headers = ["API ID", "API Name", "Method", "Endpoint", "Request Payload", "Response Payload", "Business Rule", "API Type"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    for row_idx, spec in enumerate(brd_project.api_specifications, 2):
        row_data = [
            spec.api_id or "",
            spec.api_name or "",
            spec.method or "POST",
            spec.endpoint or "",
            spec.request_payload or "",
            spec.response_payload or "",
            spec.business_rule or "",
            spec.api_type or ""
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 15


def create_llm_prompt_sheet(wb, brd_project, header_fill, header_font, border):
    """Create LLM Prompts sheet."""
    ws = wb.create_sheet("4. LLM Prompts", 3)
    
    # Headers
    headers = ["Prompt ID", "Use Case", "Model", "Temperature", "Input Variables", "Prompt Template", "Expected Output"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    for row_idx, prompt in enumerate(brd_project.llm_prompts, 2):
        row_data = [
            prompt.prompt_id or "",
            prompt.use_case or "",
            prompt.model_name or "llama3.2",
            prompt.temperature or 0.7,
            prompt.input_variables or "",
            prompt.prompt_template or "",
            prompt.expected_output or ""
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 30


def create_database_schema_sheet(wb, brd_project, header_fill, header_font, border):
    """Create Database Schema sheet."""
    ws = wb.create_sheet("5. Database Schema", 4)
    
    # Headers
    headers = ["Table Name", "Field Name", "Data Type", "Constraints", "Relationship", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    for row_idx, field in enumerate(brd_project.database_schema, 2):
        row_data = [
            field.table_name or "",
            field.field_name or "",
            field.data_type or "VARCHAR",
            field.constraints or "",
            field.relationship or "N/A",
            field.description or ""
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30


def create_tech_stack_sheet(wb, brd_project, header_fill, header_font, border):
    """Create Tech Stack & Version Control sheet."""
    ws = wb.create_sheet("6. Tech Stack & VC", 5)
    
    # Headers
    headers = ["Category", "Technology/Tool", "Version", "Rationale", "Repository URL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    for row_idx, tech in enumerate(brd_project.tech_stack, 2):
        row_data = [
            tech.category or "",
            tech.technology_tool or "",
            tech.version or "",
            tech.rationale or "",
            tech.repository_url or ""
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30


def create_traceability_sheet(wb, brd_project, header_fill, header_font, border):
    """Create Traceability Matrix sheet."""
    ws = wb.create_sheet("7. Traceability Matrix", 6)
    
    # Headers
    headers = ["Requirement ID", "Business Requirement", "Linked UI IDs", "Linked API IDs", "Linked LLM IDs", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    for row_idx, link in enumerate(brd_project.traceability_matrix, 2):
        row_data = [
            link.business_requirement_id or "",
            link.business_requirement or "",
            link.linked_ui_id or "",
            link.linked_api_id or "",
            link.linked_llm_id or "",
            link.status or "Proposed"
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15


def create_agent_architecture_sheet(wb, brd_project, header_fill, header_font, border):
    """Create Agent Architecture sheet."""
    ws = wb.create_sheet("8. Agent Architecture", 7)
    
    # Headers
    headers = ["Agent ID", "Agent Name", "Agent Type", "Primary Role", "Capabilities", "Dependencies", "Communication Protocol", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    for row_idx, agent in enumerate(brd_project.agent_architectures, 2):
        row_data = [
            agent.agent_id or "",
            agent.agent_name or "",
            agent.agent_type or "",
            agent.primary_role or "",
            agent.capabilities or "",
            agent.dependencies or "",
            agent.communication_protocol or "REST",
            agent.description or ""
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30


def create_agent_configuration_sheet(wb, brd_project, header_fill, header_font, border):
    """Create Agent Configuration sheet."""
    ws = wb.create_sheet("9. Agent Configuration", 8)
    
    # Headers
    headers = ["Config ID", "Agent ID", "Parameter Name", "Parameter Value", "Parameter Type", "Description", "Required"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    for row_idx, config in enumerate(brd_project.agent_configurations, 2):
        row_data = [
            config.config_id or "",
            config.agent_id or "",
            config.parameter_name or "",
            config.parameter_value or "",
            config.parameter_type or "string",
            config.description or "",
            "Yes" if config.required else "No"
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 10


def create_agent_task_sheet(wb, brd_project, header_fill, header_font, border):
    """Create Agent Task sheet."""
    ws = wb.create_sheet("10. Agent Tasks", 9)
    
    # Headers
    headers = ["Task ID", "Agent ID", "Task Name", "Task Type", "Input Data", "Output Data", "Success Criteria", "Error Handling", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    for row_idx, task in enumerate(brd_project.agent_tasks, 2):
        row_data = [
            task.task_id or "",
            task.agent_id or "",
            task.task_name or "",
            task.task_type or "",
            task.input_data or "",
            task.output_data or "",
            task.success_criteria or "",
            task.error_handling or "",
            task.description or ""
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 30
